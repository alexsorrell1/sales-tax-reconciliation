"""Multistate Sales & Use Tax Reconciliation Engine
====================================================

Reads a raw payment-processor transaction export, a tax rate reference
table, and the remittance (filing) history, then uses SQL to find every
discrepancy an accountant would need to explain before signing off:

    1. DUPLICATE_POSTING        the processor posted the same transaction twice
    2. MISSING_TAX              taxable orders where zero tax was charged
    3. RATE_MISMATCH            tax charged at the wrong rate
    4. Remittance gaps          filed/paid amounts that don't match collections

It produces three deliverables in the output/ folder:

    * tax_reconciliation_report.xlsx  formatted, color-coded workbook
    * tax_reconciliation_charts.png   variance by state + monthly trend
    * executive_summary.txt           auto-written narrative for leadership

Run it from the project folder with:  python tax_reconciler.py
No arguments needed. All logic lives in the sql/ folder so the SQL can
be read, reviewed, and reused on its own.
"""

import sqlite3
import sys
import textwrap
from pathlib import Path

import pandas as pd
import matplotlib

matplotlib.use("Agg")  # draw charts to a file, no screen needed
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------
# Folder locations, always relative to this file so the script
# works no matter where you run it from.
# ---------------------------------------------------------------
PROJECT_DIR = Path(__file__).resolve().parent
DATA_DIR = PROJECT_DIR / "data"
SQL_DIR = PROJECT_DIR / "sql"
OUTPUT_DIR = PROJECT_DIR / "output"

MATERIALITY = 25.00  # dollars; state-months beyond this get flagged REVIEW

# Colors for the Excel report
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")      # exceptions
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")    # informational
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")    # clean
MONEY_FMT = "#,##0.00;[Red](#,##0.00)"


def load_database():
    """Create an in-memory SQLite database and load the three CSVs."""
    for name in ("transactions.csv", "tax_rates.csv", "remittances.csv"):
        if not (DATA_DIR / name).exists():
            sys.exit(
                f"ERROR: data/{name} not found. Make sure the data folder "
                "sits next to tax_reconciler.py and contains all three CSVs."
            )

    conn = sqlite3.connect(":memory:")
    conn.executescript((SQL_DIR / "01_create_tables.sql").read_text())

    for table in ("transactions", "tax_rates", "remittances"):
        df = pd.read_csv(DATA_DIR / f"{table}.csv")
        df.to_sql(table, conn, if_exists="append", index=False)

    return conn


def run_reconciliation(conn):
    """Execute the two analysis queries and return their results."""
    exceptions = pd.read_sql_query(
        (SQL_DIR / "02_detect_exceptions.sql").read_text(), conn
    )
    recon = pd.read_sql_query(
        (SQL_DIR / "03_reconcile_remittances.sql").read_text(), conn
    )
    return exceptions, recon


def summarize(exceptions, recon):
    """Boil both result sets down to the numbers leadership asks about."""
    exc = exceptions[exceptions["exception_flag"] != "OK"]

    undercollected = exc.loc[
        exc["exception_flag"].isin(["RATE_MISMATCH", "MISSING_TAX"]),
        "variance_impact",
    ].sum()
    duplicates = exc.loc[
        exc["exception_flag"] == "DUPLICATE_POSTING", "variance_impact"
    ].sum()
    remit_gap = recon.loc[recon["remittance_variance"].abs() > 0.02,
                          "remittance_variance"].sum()

    counts = exc["exception_flag"].value_counts().to_dict()

    worst = (
        exc[exc["exception_flag"].isin(["RATE_MISMATCH", "MISSING_TAX"])]
        .groupby("state")["variance_impact"]
        .sum()
        .sort_values()
    )
    worst_state = worst.index[0] if len(worst) else "n/a"
    worst_amount = worst.iloc[0] if len(worst) else 0.0

    return {
        "undercollected": undercollected,
        "duplicates": duplicates,
        "remit_gap": remit_gap,
        "counts": counts,
        "worst_state": worst_state,
        "worst_amount": worst_amount,
        "total_exposure": abs(undercollected) + abs(remit_gap),
        "review_periods": int((recon["status"] == "REVIEW").sum()),
    }


def style_sheet(ws, money_cols, status_col=None, flag_col=None):
    """Apply the standard formatting to a worksheet."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for col_idx in money_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = MONEY_FMT

    for col in range(1, ws.max_column + 1):
        width = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, min(ws.max_row, 200) + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(width + 3, 28)

    for row in range(2, ws.max_row + 1):
        if status_col:
            val = ws.cell(row=row, column=status_col).value
            fill = RED_FILL if val == "REVIEW" else GREEN_FILL
            ws.cell(row=row, column=status_col).fill = fill
        if flag_col:
            val = ws.cell(row=row, column=flag_col).value
            if val in ("MISSING_TAX", "RATE_MISMATCH", "DUPLICATE_POSTING"):
                ws.cell(row=row, column=flag_col).fill = RED_FILL
            elif val == "MARKETPLACE_FACILITATOR":
                ws.cell(row=row, column=flag_col).fill = AMBER_FILL


def write_excel(exceptions, recon, stats):
    """Build the formatted three-tab workbook."""
    path = OUTPUT_DIR / "tax_reconciliation_report.xlsx"
    exc_detail = exceptions[exceptions["exception_flag"] != "OK"].copy()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        recon.to_excel(writer, sheet_name="State-Month Reconciliation",
                       index=False)
        exc_detail.to_excel(writer, sheet_name="Exception Detail", index=False)

        readme = pd.DataFrame(
            {
                "Field": [
                    "What this workbook is",
                    "Tab 1: State-Month Reconciliation",
                    "Tab 2: Exception Detail",
                    "Materiality threshold",
                    "Total undercollected tax",
                    "Duplicate postings (gross overstated)",
                    "Remittance gap",
                    "Generated by",
                ],
                "Value": [
                    "Three-way sales & use tax reconciliation: expected vs collected vs remitted",
                    "Deduped direct sales by state and filing period, REVIEW flags beyond materiality",
                    "Every flagged transaction with its dollar impact",
                    f"${MATERIALITY:,.2f} per state-month",
                    f"${abs(stats['undercollected']):,.2f}",
                    f"${stats['duplicates']:,.2f}",
                    f"${abs(stats['remit_gap']):,.2f}",
                    "tax_reconciler.py (SQL logic in the sql/ folder)",
                ],
            }
        )
        readme.to_excel(writer, sheet_name="Read Me", index=False)

        wb = writer.book
        recon_cols = list(recon.columns)
        style_sheet(
            wb["State-Month Reconciliation"],
            money_cols=[recon_cols.index(c) + 1 for c in (
                "taxable_sales", "expected_tax", "tax_collected",
                "tax_remitted", "collection_variance", "remittance_variance",
                "cumulative_collection_variance")],
            status_col=recon_cols.index("status") + 1,
        )
        exc_cols = list(exc_detail.columns)
        style_sheet(
            wb["Exception Detail"],
            money_cols=[exc_cols.index(c) + 1 for c in (
                "taxable_sales", "tax_collected", "expected_tax",
                "collection_variance", "variance_impact")],
            flag_col=exc_cols.index("exception_flag") + 1,
        )
        style_sheet(wb["Read Me"], money_cols=[])

    return path


def write_charts(exceptions, recon):
    """Two charts: undercollection by state, and the monthly trend."""
    path = OUTPUT_DIR / "tax_reconciliation_charts.png"
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))

    by_state = (
        recon.groupby("state")["collection_variance"].sum().sort_values()
    )
    colors = ["#C00000" if v < 0 else "#2E7D32" for v in by_state]
    ax1.bar(by_state.index, by_state.values, color=colors)
    ax1.axhline(0, color="black", linewidth=0.8)
    ax1.margins(y=0.20)
    ax1.set_title("Collection Variance by State\n(collected minus expected, direct sales)")
    ax1.set_ylabel("Dollars")
    for i, v in enumerate(by_state.values):
        if abs(v) < 0.01:
            continue  # don't label zero bars
        offset, va = (4, "bottom") if v > 0 else (-4, "top")
        ax1.annotate(f"{v:,.0f}", (i, v), textcoords="offset points",
                     xytext=(0, offset), ha="center", va=va, fontsize=9)

    by_month = recon.groupby("month")[
        ["collection_variance", "remittance_variance"]].sum()
    ax2.plot(by_month.index, by_month["collection_variance"],
             marker="o", label="Collection variance")
    ax2.plot(by_month.index, by_month["remittance_variance"],
             marker="s", label="Remittance variance")
    ax2.axhline(0, color="black", linewidth=0.8)
    ax2.set_title("Monthly Variance Trend")
    ax2.set_ylabel("Dollars")
    ax2.legend()
    ax2.tick_params(axis="x", rotation=45)

    for ax in (ax1, ax2):
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    fig.suptitle("Multistate Sales & Use Tax Reconciliation", fontsize=14,
                 fontweight="bold")
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)
    return path


def write_summary(stats, recon):
    """Auto-write the plain-English summary a controller actually reads."""
    path = OUTPUT_DIR / "executive_summary.txt"
    months = sorted(recon["month"].unique())
    c = stats["counts"]

    def para(text, indent="   "):
        """Wrap a finding into tidy 68-character lines."""
        return textwrap.fill(" ".join(text.split()), width=68,
                             subsequent_indent=indent)

    findings = [
        para(
            f"1. Undercollected tax of ${abs(stats['undercollected']):,.2f} "
            f"across {c.get('RATE_MISMATCH', 0)} wrong-rate and "
            f"{c.get('MISSING_TAX', 0)} zero-tax transactions. The wrong-rate "
            "cluster traces to the processor applying a state-only rate and "
            "dropping local/district taxes; the zero-tax orders indicate a "
            "nexus configuration error. Largest state impact: "
            f"{stats['worst_state']} at ${abs(stats['worst_amount']):,.2f}."
        ),
        para(
            "2. Duplicate postings overstated gross collected tax by "
            f"${stats['duplicates']:,.2f} ({c.get('DUPLICATE_POSTING', 0)} "
            "re-posted transactions). Filings prepared off the raw processor "
            "report would overstate liability; the dedup step removes them."
        ),
        para(
            f"3. Remittance gap of ${abs(stats['remit_gap']):,.2f}: tax "
            "collected from customers but not yet filed/paid for at least "
            "one state-period. Collected-but-unremitted tax is a trust-fund "
            "liability and should be trued up in the next filing cycle."
        ),
        para(
            "4. Marketplace-facilitator sales "
            f"({c.get('MARKETPLACE_FACILITATOR', 0)} transactions) were "
            "correctly excluded from company liability; the facilitator "
            "remits tax on those orders."
        ),
    ]

    actions = [
        para(
            "1. Correct the processor's rate table for the affected state "
            "and backfile amended returns for the undercollected periods."
        ),
        para(
            "2. Fix the nexus configuration causing zero-tax checkouts and "
            "add a monthly zero-tax exception report as a preventive control."
        ),
        para(
            "3. Add this dedup check to the close checklist so filings are "
            "always prepared from deduplicated data."
        ),
        para(
            "4. Remit the outstanding collected-not-remitted balance with "
            "the next filing and reconcile remittances to collections monthly."
        ),
    ]

    lines = [
        "EXECUTIVE SUMMARY - MULTISTATE SALES & USE TAX RECONCILIATION",
        f"Periods covered: {months[0]} through {months[-1]}",
        "=" * 68,
        "",
        f"Total estimated exposure: ${stats['total_exposure']:,.2f}",
        f"State-month periods flagged for review: {stats['review_periods']}",
        "",
        "KEY FINDINGS",
        "-" * 68,
        findings[0], "", findings[1], "", findings[2], "", findings[3],
        "",
        "RECOMMENDED ACTIONS",
        "-" * 68,
        *actions,
        "",
        para(
            "Prepared automatically by tax_reconciler.py; transaction-level "
            "support is in tax_reconciliation_report.xlsx (Exception Detail "
            "tab).", indent=""
        ),
    ]
    path.write_text("\n".join(lines))
    return path


def main():
    print("Multistate Sales & Use Tax Reconciliation")
    print("-" * 45)
    OUTPUT_DIR.mkdir(exist_ok=True)

    conn = load_database()
    print("1/5  Data loaded into SQLite")

    exceptions, recon = run_reconciliation(conn)
    conn.close()
    n_exc = int((exceptions["exception_flag"] != "OK").sum())
    print(f"2/5  SQL reconciliation complete: {len(exceptions)} rows analyzed, "
          f"{n_exc} exceptions flagged")

    stats = summarize(exceptions, recon)
    xlsx = write_excel(exceptions, recon, stats)
    print(f"3/5  Excel report written: {xlsx.name}")

    png = write_charts(exceptions, recon)
    print(f"4/5  Charts written: {png.name}")

    txt = write_summary(stats, recon)
    print(f"5/5  Executive summary written: {txt.name}")

    print("-" * 45)
    print(f"Total estimated exposure: ${stats['total_exposure']:,.2f}")
    print(f"Done. All deliverables are in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
